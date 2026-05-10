import os, io, json, sqlite3, datetime, copy
import pandas as pd
import plotly.graph_objects as go
import dash
from dash import dcc, html, Input, Output, State, ctx
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
DB   = os.path.join(BASE, "..", "frota_motos_completo.sqlite")
GEO  = os.path.join(BASE, "assets", "br_states.geojson")

# ── Design tokens ──────────────────────────────────────────────────────────────
BG     = "#f5f8fa"   # page background
CARD   = "#ffffff"   # card surface
SIDE   = "#f0f5f8"   # sidebar tint
BORDER = "#e2eef5"   # hairline border
DEEP   = "#1b4965"   # Deep Sea — navy, primary text, featured card bg
STEEL  = "#5fa8d3"   # Steel Blue
CYAN   = "#62b6cb"   # Cyan
MIST   = "#bee9e8"   # Mist teal
SKY    = "#cae9ff"   # Sky
TEXT   = "#1b4965"   # body text
LABEL  = "#94b8cc"   # muted label / axis text
WARN   = "#b25e1c"   # amber
DANGER = "#be2f2f"   # red
GRID   = "#e8f2f8"   # gridlines

CHART_THEME = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor ="rgba(0,0,0,0)",
    font=dict(family="Montserrat, sans-serif", color=LABEL, size=11),
    hoverlabel=dict(bgcolor=DEEP, font_color="#ffffff", bordercolor=STEEL, font_size=12),
    margin=dict(l=4, r=16, t=8, b=4),
)

# ── State mapping ──────────────────────────────────────────────────────────────
UF_SIGLA = {
    "ACRE":"AC","ALAGOAS":"AL","AMAPA":"AP","AMAZONAS":"AM","BAHIA":"BA",
    "CEARA":"CE","DISTRITO FEDERAL":"DF","ESPIRITO SANTO":"ES","GOIAS":"GO",
    "MARANHAO":"MA","MATO GROSSO":"MT","MATO GROSSO DO SUL":"MS",
    "MINAS GERAIS":"MG","PARA":"PA","PARAIBA":"PB","PARANA":"PR",
    "PERNAMBUCO":"PE","PIAUI":"PI","RIO DE JANEIRO":"RJ",
    "RIO GRANDE DO NORTE":"RN","RIO GRANDE DO SUL":"RS","RONDONIA":"RO",
    "RORAIMA":"RR","SANTA CATARINA":"SC","SAO PAULO":"SP",
    "SERGIPE":"SE","TOCANTINS":"TO",
}

# ── Helpers ────────────────────────────────────────────────────────────────────
def conn():
    c = sqlite3.connect(DB)
    c.execute("PRAGMA cache_size = -32000")
    c.execute("PRAGMA temp_store  = MEMORY")
    c.execute("PRAGMA mmap_size   = 268435456")
    return c

def fmt(n):
    if n is None or (isinstance(n, float) and n != n): return "—"
    n = int(n)
    if n >= 1_000_000: return f"{n/1_000_000:.2f}M"
    if n >= 1_000:     return f"{n/1_000:.1f}K"
    return f"{n:,}"

def where(fab, mod, uf, mun, ano_ini, ano_fim, modalidade=None):
    c = [f"Ano_Fabricacao >= {ano_ini}", f"Ano_Fabricacao <= {ano_fim}"]
    if modalidade:
        lst = "','".join(modalidade)
        c.append(f"Modalidade IN ('{lst}')")
    if uf:
        lst = "','".join(uf)
        c.append(f"UF IN ('{lst}')")
    if mun:
        lst = "','".join(mun)
        c.append(f"Municipio IN ('{lst}')")
    if fab:
        lst = "','".join(fab)
        c.append(f"Fabricante IN ('{lst}')")
    if mod:
        lst = "','".join(mod)
        c.append(f"Marca_Modelo IN ('{lst}')")
    return " AND ".join(c)

def q(sql):
    c = conn()
    df = pd.read_sql(sql, c)
    c.close()
    return df

# ── Static data ────────────────────────────────────────────────────────────────
fab_options_df = q("""
    SELECT Fabricante AS fab, SUM(Qtd_Veiculos) AS total
    FROM frota WHERE Fabricante IS NOT NULL AND Fabricante != ''
    GROUP BY Fabricante ORDER BY total DESC
""")
uf_list = q("SELECT DISTINCT UF FROM frota ORDER BY UF")["UF"].tolist()

YEAR_OPTS = [{"label": str(y), "value": y} for y in range(1970, 2027)]

MODAL_OPTS = [
    {"label": "🏁  Sport",                       "value": "Sport"},
    {"label": "🔥  Naked / Roadster",             "value": "Naked/Roadster"},
    {"label": "🌍  Adventure / Trail",            "value": "Adventure/Trail"},
    {"label": "🛠  Custom / Classic / Scrambler",  "value": "Custom/Classic/Scrambler"},
    {"label": "🏔  Off-road / Motocross / Enduro", "value": "Off-road/Motocross/Enduro"},
    {"label": "🛵  Scooter",                      "value": "Scooter"},
    {"label": "📦  Outros",                       "value": "Outros"},
]

with open(GEO) as f:
    geojson = json.load(f)

FAB_OPTS = [{"label": f"{r.fab}  ({fmt(r.total)})", "value": r.fab}
            for r in fab_options_df.itertuples()]
UF_OPTS  = [{"label": uf.title(), "value": uf} for uf in uf_list]

# ── UI primitives ──────────────────────────────────────────────────────────────
_DD = {"fontSize": "0.82rem"}

def kpi_card(label_text, vid, featured=False):
    bg     = DEEP if featured else CARD
    brd    = f"1px solid {DEEP}" if featured else f"1px solid {BORDER}"
    lbl_c  = MIST   if featured else LABEL
    val_c  = "#ffffff" if featured else DEEP
    return html.Div([
        html.P(label_text, style={
            "margin": "0 0 18px 0",
            "fontSize": "0.6rem", "fontWeight": "600",
            "letterSpacing": "0.14em", "textTransform": "uppercase",
            "color": lbl_c,
        }),
        html.Div(id=vid, className="kpi-value", style={
            "fontSize": "2.6rem", "fontWeight": "700",
            "color": val_c, "letterSpacing": "-0.04em", "lineHeight": "1",
        }),
    ], style={
        "background": bg, "border": brd, "borderRadius": "14px",
        "padding": "1.5rem 1.6rem", "flex": "1", "minWidth": "140px",
    })

def chart_card(title, child_id, height, flex=1, extra_style=None, expandable=False):
    s = {
        "background": CARD, "border": f"1px solid {BORDER}",
        "borderRadius": "14px", "padding": "1.4rem 1.6rem", "flex": flex,
    }
    if extra_style:
        s.update(extra_style)
    _btn = {
        "background": "transparent", "border": f"1px solid {BORDER}",
        "color": LABEL, "cursor": "pointer", "borderRadius": "6px",
        "padding": "3px 9px", "fontSize": "0.85rem", "lineHeight": "1.5",
        "transition": "all 0.18s",
    }
    header = [
        html.P(title, style={
            "margin": 0, "fontSize": "0.6rem", "fontWeight": "600",
            "letterSpacing": "0.14em", "textTransform": "uppercase", "color": LABEL,
        }),
    ]
    if expandable:
        header.append(html.Div([
            html.Button("↓", id=f"dl-{child_id}", n_clicks=0,
                title="Baixar como Excel", className="download-btn", style=_btn),
            html.Button("⛶", id=f"expand-{child_id}", n_clicks=0,
                title="Expandir", className="expand-btn", style=_btn),
        ], style={"display": "flex", "gap": "6px"}))
    return html.Div([
        html.Div(header, style={
            "display": "flex", "justifyContent": "space-between",
            "alignItems": "center", "marginBottom": "14px",
        }),
        dcc.Graph(id=child_id, config={"displayModeBar": False, "responsive": True},
                  style={"height": f"{height}px"}),
    ], style=s, className="chart-card")

def sec_label(txt):
    return html.P(txt, style={
        "margin": "0 0 7px 0", "fontSize": "0.6rem", "fontWeight": "600",
        "letterSpacing": "0.13em", "textTransform": "uppercase", "color": LABEL,
    })

def sidebar_sec(title, content):
    return html.Div([sec_label(title), content], style={"marginBottom": "1.3rem"})

def sub_label(txt):
    return html.P(txt, style={
        "margin": "0 0 5px 0", "fontSize": "0.58rem", "fontWeight": "600",
        "letterSpacing": "0.1em", "textTransform": "uppercase", "color": LABEL,
    })

# ── App ────────────────────────────────────────────────────────────────────────
app = dash.Dash(
    __name__,
    title="Frota Motos · Brasil",
    suppress_callback_exceptions=True,
)

app.layout = html.Div([

    # ── Topbar ────────────────────────────────────────────────────────────────
    html.Div([
        html.Div([
            html.P("SENATRAN · 1970 – 2026 · SÉRIE HISTÓRICA", style={
                "margin": "0 0 4px 0", "fontSize": "0.58rem", "fontWeight": "600",
                "letterSpacing": "0.14em", "color": LABEL,
            }),
            html.H1("Frota de Motos · Brasil", style={
                "margin": 0, "fontSize": "1.45rem", "fontWeight": "700",
                "color": TEXT, "letterSpacing": "-0.025em", "lineHeight": "1",
            }),
        ]),
        html.Div([
            html.Div(id="active-tag", style={
                "fontSize": "0.68rem", "color": TEXT,
                "background": "#f0f7fb", "border": f"1px solid {BORDER}",
                "borderRadius": "20px", "padding": "5px 16px",
            }),
            html.Button("↺  Limpar filtros", id="btn-reset", n_clicks=0,
                className="reset-btn",
                style={
                    "background": "transparent", "border": f"1px solid {DANGER}",
                    "borderRadius": "20px", "color": DANGER, "cursor": "pointer",
                    "fontSize": "0.7rem", "fontWeight": "600",
                    "letterSpacing": "0.04em", "padding": "5px 16px",
                    "transition": "all 0.18s", "whiteSpace": "nowrap",
                }),
        ], style={"display": "flex", "alignItems": "center", "gap": "10px"}),
    ], style={
        "display": "flex", "alignItems": "center", "justifyContent": "space-between",
        "padding": "1rem 2rem",
        "background": CARD,
        "borderBottom": f"1px solid {BORDER}",
        "position": "sticky", "top": 0, "zIndex": 200,
    }),

    # ── Body ──────────────────────────────────────────────────────────────────
    html.Div([

        # ── Sidebar ───────────────────────────────────────────────────────────
        html.Div([

            sidebar_sec("Modalidade", dcc.Dropdown(
                id="f-modal", options=MODAL_OPTS, value=[], multi=True,
                placeholder="Todas as modalidades…", style=_DD,
            )),
            sidebar_sec("Fabricante", dcc.Dropdown(
                id="f-fab", options=FAB_OPTS, value=None, multi=True,
                placeholder="Todas as marcas…", style=_DD,
            )),
            sidebar_sec("Modelo", dcc.Dropdown(
                id="f-mod", options=[], value=[], multi=True,
                placeholder="Selecione um fabricante…",
                disabled=True, style=_DD,
            )),
            sidebar_sec("Estado", dcc.Dropdown(
                id="f-uf", options=UF_OPTS, value=[], multi=True,
                placeholder="Todos os estados…", style=_DD,
            )),
            sidebar_sec("Município", dcc.Dropdown(
                id="f-mun", options=[], value=[], multi=True,
                placeholder="Selecione um estado…",
                disabled=True, style=_DD,
            )),

            # Ano de fabricação
            html.Div([
                sec_label("Ano de fabricação"),
                html.Div([
                    html.Div([
                        sub_label("De"),
                        dcc.Dropdown(
                            id="f-ano-ini", options=YEAR_OPTS, value=1970,
                            clearable=False, style=_DD,
                        ),
                    ], style={"flex": 1}),
                    html.Div([
                        sub_label("Até"),
                        dcc.Dropdown(
                            id="f-ano-fim", options=YEAR_OPTS, value=2026,
                            clearable=False, style=_DD,
                        ),
                    ], style={"flex": 1}),
                ], style={"display": "flex", "gap": "8px"}),
            ], style={"marginBottom": "1.3rem"}),

            html.Hr(style={"border": "none", "borderTop": f"1px solid {BORDER}", "margin": "0.5rem 0 1rem"}),

            html.P("Clique num estado no mapa para filtrar.", style={
                "fontSize": "0.65rem", "color": LABEL, "lineHeight": "1.5", "margin": 0,
            }),

        ], style={
            "width": "255px", "flexShrink": 0,
            "padding": "1.5rem 1.2rem",
            "background": SIDE,
            "borderRight": f"1px solid {BORDER}",
            "height": "calc(100vh - 61px)",
            "overflowY": "auto", "position": "sticky", "top": "61px",
        }),

        # ── Main ──────────────────────────────────────────────────────────────
        html.Div([

            # KPI row
            html.Div([
                kpi_card("Total de Motos",    "k-total",   featured=True),
                kpi_card("Modelos Distintos", "k-modelos"),
                kpi_card("Municípios",        "k-munis"),
                kpi_card("Estados",           "k-estados"),
                kpi_card("Média por Modelo",  "k-media"),
            ], style={
                "display": "flex", "gap": "10px",
                "marginBottom": "14px", "flexWrap": "wrap",
            }),

            # Map + Fabricantes
            html.Div([
                chart_card("Distribuição geográfica — clique para filtrar",
                           "g-mapa", 390, flex=3),
                chart_card("Top fabricantes", "g-fab", 390, flex=2, expandable=True),
            ], style={"display": "flex", "gap": "12px", "marginBottom": "12px"}),

            # Top Modelos + Evolução
            html.Div([
                chart_card("Top 15 modelos",   "g-modelos", 430, flex=3, expandable=True),
                chart_card("Evolução por ano", "g-anos",    430, flex=2, expandable=True),
            ], style={"display": "flex", "gap": "12px", "marginBottom": "12px"}),

            # Municípios
            chart_card("Top 20 municípios", "g-munis", 310,
                       extra_style={"marginBottom": "12px"}, expandable=True),

        ], style={
            "flex": 1, "padding": "1.4rem",
            "overflowY": "auto", "height": "calc(100vh - 61px)",
        }),

    ], style={"display": "flex", "height": "calc(100vh - 61px)"}),

    # ── Download trigger ──────────────────────────────────────────────────────
    dcc.Download(id="dl-chart"),

    # ── Modal overlay ─────────────────────────────────────────────────────────
    html.Div([
        html.Div([
            html.Div([
                html.P(id="modal-title", style={
                    "margin": 0, "fontSize": "0.6rem", "fontWeight": "600",
                    "letterSpacing": "0.14em", "textTransform": "uppercase",
                    "color": LABEL,
                }),
                html.Button("✕", id="modal-close", n_clicks=0,
                    style={
                        "background": "transparent", "border": f"1px solid {BORDER}",
                        "color": LABEL, "cursor": "pointer", "borderRadius": "8px",
                        "padding": "4px 12px", "fontSize": "0.9rem", "fontWeight": "600",
                        "lineHeight": "1", "transition": "all 0.18s",
                    }),
            ], style={"display": "flex", "justifyContent": "space-between",
                      "alignItems": "center", "marginBottom": "16px"}),
            dcc.Graph(id="modal-graph",
                      config={"displayModeBar": True, "responsive": True},
                      style={"height": "calc(100vh - 160px)"}),
        ], style={
            "background": CARD, "borderRadius": "16px", "padding": "1.6rem 2rem",
            "width": "calc(100vw - 80px)", "height": "calc(100vh - 80px)",
            "maxWidth": "1600px", "border": f"1px solid {BORDER}",
        }),
    ], id="modal-overlay", style={
        "display": "none", "position": "fixed", "top": 0, "left": 0,
        "width": "100vw", "height": "100vh",
        "background": "rgba(27,73,101,0.6)",
        "zIndex": 999, "justifyContent": "center", "alignItems": "center",
        "backdropFilter": "blur(8px)",
    }),

], style={
    "background": BG, "minHeight": "100vh",
    "fontFamily": "'Montserrat', -apple-system, sans-serif",
    "color": TEXT,
})

# ── Cascading: modelos por fabricante ──────────────────────────────────────────
@app.callback(
    Output("f-mod", "options"),
    Output("f-mod", "value"),
    Output("f-mod", "disabled"),
    Input("f-fab", "value"),
)
def update_mod_opts(fab):
    if not fab:
        return [], [], True
    lst = "','".join(fab)
    df = q(f"""
        SELECT DISTINCT Marca_Modelo FROM frota
        WHERE Fabricante IN ('{lst}')
        ORDER BY Marca_Modelo
    """)
    opts = [{"label": m, "value": m} for m in df["Marca_Modelo"]]
    return opts, [], False

# ── Cascading: municípios por estado ──────────────────────────────────────────
@app.callback(
    Output("f-mun", "options"),
    Output("f-mun", "value"),
    Output("f-mun", "disabled"),
    Input("f-uf", "value"),
)
def update_mun_opts(uf):
    if not uf:
        return [], [], True
    lst = "','".join(uf)
    df = q(f"""
        SELECT DISTINCT Municipio FROM frota
        WHERE UF IN ('{lst}')
        ORDER BY Municipio
    """)
    opts = [{"label": m.title(), "value": m} for m in df["Municipio"]]
    return opts, [], False

# ── Map click → UF filter ──────────────────────────────────────────────────────
@app.callback(
    Output("f-uf", "value"),
    Input("g-mapa", "clickData"),
    prevent_initial_call=True,
)
def map_click(click):
    if not click:
        return []
    sigla = click["points"][0].get("location")
    if not sigla:
        return []
    SIGLA_UF = {v: k for k, v in UF_SIGLA.items()}
    uf = SIGLA_UF.get(sigla)
    return [uf] if uf else []

# ── Main update callback ───────────────────────────────────────────────────────
@app.callback(
    Output("k-total",    "children"),
    Output("k-modelos",  "children"),
    Output("k-munis",    "children"),
    Output("k-estados",  "children"),
    Output("k-media",    "children"),
    Output("g-mapa",     "figure"),
    Output("g-fab",      "figure"),
    Output("g-modelos",  "figure"),
    Output("g-anos",     "figure"),
    Output("g-munis",    "figure"),
    Output("active-tag", "children"),
    Input("f-modal",   "value"),
    Input("f-fab",     "value"),
    Input("f-mod",     "value"),
    Input("f-uf",      "value"),
    Input("f-mun",     "value"),
    Input("f-ano-ini", "value"),
    Input("f-ano-fim", "value"),
)
def update(modalidade, fab, mod, uf, mun, ano_ini, ano_fim):
    if ano_ini is None:  ano_ini    = 1970
    if ano_fim is None:  ano_fim    = 2026
    if not uf:           uf         = []
    if not mun:          mun        = []
    if not mod:          mod        = []
    if not modalidade:   modalidade = []
    w = where(fab, mod, uf, mun, ano_ini, ano_fim, modalidade)

    db = conn()

    kpis = pd.read_sql(f"""
        SELECT SUM(Qtd_Veiculos)            AS total,
               COUNT(DISTINCT Marca_Modelo) AS modelos,
               COUNT(DISTINCT Municipio)    AS munis,
               COUNT(DISTINCT UF)           AS estados
        FROM frota WHERE {w}
    """, db).iloc[0]

    media = (kpis["total"] / kpis["modelos"]) if kpis["modelos"] else 0

    map_df = pd.read_sql(f"""
        SELECT UF, SUM(Qtd_Veiculos) AS total FROM frota
        WHERE {w} GROUP BY UF ORDER BY total DESC
    """, db)
    map_df["sigla"] = map_df["UF"].map(UF_SIGLA)
    map_df["nome"]  = map_df["UF"].str.title()

    fab_df = pd.read_sql(f"""
        SELECT Fabricante AS fab, SUM(Qtd_Veiculos) AS total
        FROM frota WHERE {w} AND Fabricante IS NOT NULL AND Fabricante != ''
        GROUP BY Fabricante ORDER BY total DESC LIMIT 12
    """, db)

    mod_df = pd.read_sql(f"""
        SELECT Marca_Modelo, SUM(Qtd_Veiculos) AS total
        FROM frota WHERE {w}
        GROUP BY Marca_Modelo ORDER BY total DESC LIMIT 15
    """, db)

    ano_df = pd.read_sql(f"""
        SELECT Ano_Fabricacao, SUM(Qtd_Veiculos) AS total
        FROM frota WHERE {w}
        GROUP BY Ano_Fabricacao ORDER BY Ano_Fabricacao
    """, db)

    mun_df = pd.read_sql(f"""
        SELECT Municipio, UF, SUM(Qtd_Veiculos) AS total
        FROM frota WHERE {w}
        GROUP BY Municipio, UF ORDER BY total DESC LIMIT 20
    """, db)
    db.close()

    mun_df["label"] = (mun_df["Municipio"].str.title() + " / " +
                       mun_df["UF"].map(UF_SIGLA).fillna(mun_df["UF"]))

    # ── Map ───────────────────────────────────────────────────────────────────
    fig_map = go.Figure(go.Choropleth(
        geojson=geojson,
        locations=map_df["sigla"].dropna(),
        z=map_df.loc[map_df["sigla"].notna(), "total"],
        featureidkey="properties.sigla",
        colorscale=[
            [0.00, "#f0f7fb"],
            [0.20, MIST],
            [0.50, CYAN],
            [0.78, STEEL],
            [1.00, DEEP],
        ],
        marker_line_color="#ffffff",
        marker_line_width=1.2,
        showscale=True,
        colorbar=dict(
            thickness=9, len=0.6, x=1.01,
            tickfont=dict(color=LABEL, size=9),
            title=dict(text="", font=dict(color=LABEL)),
            bgcolor="rgba(0,0,0,0)",
        ),
        customdata=map_df.loc[map_df["sigla"].notna(), ["nome", "total"]],
        hovertemplate="<b>%{customdata[0]}</b><br>%{customdata[1]:,.0f} motos<extra></extra>",
    ))
    fig_map.update_geos(fitbounds="locations", visible=False, bgcolor="rgba(0,0,0,0)",
                        showcoastlines=False)
    map_th = {k: v for k, v in CHART_THEME.items() if k != "margin"}
    fig_map.update_layout(**map_th, geo=dict(bgcolor="rgba(0,0,0,0)"),
                          margin=dict(l=0, r=0, t=0, b=0), height=380)

    # ── Fabricantes ───────────────────────────────────────────────────────────
    fig_fab = go.Figure(go.Bar(
        x=fab_df["total"], y=fab_df["fab"], orientation="h",
        marker=dict(color=STEEL, opacity=0.9, line=dict(width=0)),
        text=fab_df["total"].apply(fmt),
        textposition="outside",
        textfont=dict(color=LABEL, size=10),
        hovertemplate="<b>%{y}</b> — %{x:,.0f} motos<extra></extra>",
        cliponaxis=False,
    ))
    fig_fab.update_layout(
        **CHART_THEME,
        xaxis=dict(showgrid=False, showticklabels=False, zeroline=False,
                   range=[0, fab_df["total"].max() * 1.25] if not fab_df.empty else [0, 1]),
        yaxis=dict(showgrid=False, tickfont=dict(size=11, color=TEXT),
                   autorange="reversed"),
        bargap=0.3, height=380,
    )

    # ── Top modelos ───────────────────────────────────────────────────────────
    fig_mod = go.Figure(go.Bar(
        x=mod_df["total"], y=mod_df["Marca_Modelo"], orientation="h",
        marker=dict(color=CYAN, opacity=0.9, line=dict(width=0)),
        text=mod_df["total"].apply(fmt),
        textposition="outside",
        textfont=dict(color=LABEL, size=10),
        hovertemplate="<b>%{y}</b> — %{x:,.0f} motos<extra></extra>",
        cliponaxis=False,
    ))
    fig_mod.update_layout(
        **CHART_THEME,
        xaxis=dict(showgrid=False, showticklabels=False, zeroline=False,
                   range=[0, mod_df["total"].max() * 1.25] if not mod_df.empty else [0, 1]),
        yaxis=dict(showgrid=False, tickfont=dict(size=10, color=TEXT),
                   autorange="reversed"),
        bargap=0.24, height=430,
    )

    # ── Evolução por ano ──────────────────────────────────────────────────────
    fig_ano = go.Figure()
    if not ano_df.empty:
        fig_ano.add_trace(go.Scatter(
            x=ano_df["Ano_Fabricacao"], y=ano_df["total"],
            mode="lines+markers",
            line=dict(color=DEEP, width=1.5, shape="spline"),
            marker=dict(size=5, color=DEEP, line=dict(width=0)),
            hovertemplate="<b>%{x}</b> — %{y:,.0f} motos<extra></extra>",
        ))
    fig_ano.update_layout(
        **CHART_THEME,
        xaxis=dict(
            showgrid=False, zeroline=False,
            tickfont=dict(size=10, color=LABEL),
            tickmode="linear", dtick=5, tickangle=-45,
        ),
        yaxis=dict(
            showgrid=True, gridcolor=GRID, zeroline=False,
            tickfont=dict(size=10, color=LABEL),
        ),
        height=430,
    )

    # ── Municípios ────────────────────────────────────────────────────────────
    fig_mun = go.Figure(go.Bar(
        x=mun_df["label"], y=mun_df["total"],
        marker=dict(color=STEEL, opacity=0.9, line=dict(width=0)),
        text=mun_df["total"].apply(fmt),
        textposition="outside",
        textfont=dict(color=LABEL, size=9),
        hovertemplate="<b>%{x}</b> — %{y:,.0f} motos<extra></extra>",
        cliponaxis=False,
    ))
    fig_mun.update_layout(
        **CHART_THEME,
        xaxis=dict(showgrid=False, tickfont=dict(size=9, color=LABEL),
                   tickangle=-35, zeroline=False),
        yaxis=dict(showgrid=True, gridcolor=GRID, zeroline=False,
                   tickfont=dict(size=10, color=LABEL),
                   range=[0, mun_df["total"].max() * 1.18] if not mun_df.empty else [0, 1]),
        bargap=0.22, height=310,
    )

    # ── Active filter tag ─────────────────────────────────────────────────────
    parts = []
    if modalidade: parts.append(" · ".join(modalidade))
    if fab:        parts.append(" · ".join(fab))
    if mod:        parts.append(f"{len(mod)} modelo(s)")
    if uf:         parts.append(" · ".join(u.title() for u in uf))
    if mun:        parts.append(f"{len(mun)} município(s)")
    if ano_ini != 1970 or ano_fim != 2026:
        parts.append(f"{ano_ini}–{ano_fim}")
    tag = "  ·  ".join(parts) if parts else "Sem filtros activos"

    return (
        fmt(kpis["total"]),
        fmt(kpis["modelos"]),
        fmt(kpis["munis"]),
        str(int(kpis["estados"])),
        fmt(media),
        fig_map, fig_fab, fig_mod, fig_ano, fig_mun,
        tag,
    )

# ── Download as Excel ─────────────────────────────────────────────────────────
_DL_LABELS = {
    "dl-g-fab":     ("fabricantes", "Top Fabricantes"),
    "dl-g-modelos": ("modelos",     "Top Modelos"),
    "dl-g-anos":    ("anos",        "Evolução por Ano"),
    "dl-g-munis":   ("municipios",  "Top Municípios"),
}

def _build_excel(data_df: pd.DataFrame, chart_title: str,
                 modalidade, fab, mod, uf, mun, ano_ini, ano_fim) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        def _val(v, default="Todos"):
            if not v: return default
            if isinstance(v, list): return ", ".join(str(x) for x in v)
            return str(v)

        filters_rows = [
            ("Gráfico",              chart_title),
            ("Exportado em",         datetime.datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("", ""),
            ("── Filtros aplicados ──", ""),
            ("Modalidade",           _val(modalidade)),
            ("Fabricante",           _val(fab)),
            ("Modelo",               _val(mod)),
            ("Estado",               _val(uf)),
            ("Município",            _val(mun)),
            ("Ano inicial",          str(ano_ini if ano_ini else 1970)),
            ("Ano final",            str(ano_fim if ano_fim else 2026)),
        ]
        filt_df = pd.DataFrame(filters_rows, columns=["Parâmetro", "Valor"])
        filt_df.to_excel(writer, sheet_name="Filtros", index=False)

        ws_f = writer.sheets["Filtros"]
        hdr_fill   = PatternFill("solid", fgColor="1b4965")
        alt_fill_a = PatternFill("solid", fgColor="f0f5f8")
        alt_fill_b = PatternFill("solid", fgColor="ffffff")
        thin       = Side(style="thin", color="e2eef5")
        brd        = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_idx, (param, val) in enumerate(filters_rows, start=2):
            c_p = ws_f.cell(row=row_idx, column=1, value=param)
            c_v = ws_f.cell(row=row_idx, column=2, value=val)
            fill = alt_fill_a if row_idx % 2 == 0 else alt_fill_b
            is_title = row_idx == 2
            is_sep   = isinstance(param, str) and param.startswith("──")
            for cell in (c_p, c_v):
                cell.fill      = fill
                cell.border    = brd
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if is_title:
                    cell.font = Font(bold=True, color="1b4965", size=11)
                elif is_sep:
                    cell.font = Font(bold=True, color="5fa8d3", size=10)
                else:
                    cell.font = Font(color="1b4965", size=10)

        for col in (1, 2):
            h = ws_f.cell(row=1, column=col)
            h.font      = Font(bold=True, color="FFFFFF", size=11)
            h.fill      = hdr_fill
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.border    = brd
        ws_f.column_dimensions["A"].width = 22
        ws_f.column_dimensions["B"].width = 55
        ws_f.row_dimensions[1].height     = 20

        data_df.to_excel(writer, sheet_name="Dados", index=False)
        ws_d = writer.sheets["Dados"]

        for col_idx, col_name in enumerate(data_df.columns, start=1):
            h = ws_d.cell(row=1, column=col_idx)
            h.font      = Font(bold=True, color="FFFFFF", size=11)
            h.fill      = hdr_fill
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.border    = brd
            max_len = max(
                len(str(col_name)),
                data_df[col_name].astype(str).str.len().max() if len(data_df) else 0,
            )
            ws_d.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        for row_idx in range(2, len(data_df) + 2):
            fill = alt_fill_a if row_idx % 2 == 0 else alt_fill_b
            for col_idx in range(1, len(data_df.columns) + 1):
                cell = ws_d.cell(row=row_idx, column=col_idx)
                cell.fill      = fill
                cell.font      = Font(color="1b4965", size=10)
                cell.border    = brd
                cell.alignment = Alignment(vertical="center")
        ws_d.freeze_panes = "A2"

    buf.seek(0)
    return buf.read()


@app.callback(
    Output("dl-chart", "data"),
    Input("dl-g-fab",     "n_clicks"),
    Input("dl-g-modelos", "n_clicks"),
    Input("dl-g-anos",    "n_clicks"),
    Input("dl-g-munis",   "n_clicks"),
    State("f-modal",   "value"),
    State("f-fab",     "value"),
    State("f-mod",     "value"),
    State("f-uf",      "value"),
    State("f-mun",     "value"),
    State("f-ano-ini", "value"),
    State("f-ano-fim", "value"),
    prevent_initial_call=True,
)
def download_chart(n1, n2, n3, n4,
                   modalidade, fab, mod, uf, mun, ano_ini, ano_fim):
    tid = ctx.triggered_id
    if not tid or not tid.startswith("dl-"):
        return dash.no_update

    slug, chart_title = _DL_LABELS[tid]

    if ano_ini is None: ano_ini    = 1970
    if ano_fim is None: ano_fim    = 2026
    if not uf:          uf         = []
    if not mun:         mun        = []
    if not mod:         mod        = []
    if not modalidade:  modalidade = []

    w  = where(fab, mod, uf, mun, ano_ini, ano_fim, modalidade)
    db = conn()

    if tid == "dl-g-fab":
        df = pd.read_sql(f"""
            SELECT Fabricante,
                   SUM(Qtd_Veiculos) AS "Total de Motos"
            FROM frota WHERE {w}
              AND Fabricante IS NOT NULL AND Fabricante != ''
            GROUP BY Fabricante ORDER BY "Total de Motos" DESC
        """, db)
    elif tid == "dl-g-modelos":
        df = pd.read_sql(f"""
            SELECT Marca_Modelo AS "Marca / Modelo",
                   SUM(Qtd_Veiculos) AS "Total de Motos"
            FROM frota WHERE {w}
            GROUP BY Marca_Modelo ORDER BY "Total de Motos" DESC
        """, db)
    elif tid == "dl-g-anos":
        df = pd.read_sql(f"""
            SELECT Ano_Fabricacao AS "Ano de Fabricação",
                   SUM(Qtd_Veiculos) AS "Total de Motos"
            FROM frota WHERE {w}
            GROUP BY Ano_Fabricacao ORDER BY Ano_Fabricacao
        """, db)
    else:
        df = pd.read_sql(f"""
            SELECT Municipio AS Município,
                   UF,
                   SUM(Qtd_Veiculos) AS "Total de Motos"
            FROM frota WHERE {w}
            GROUP BY Municipio, UF ORDER BY "Total de Motos" DESC
        """, db)
        df["UF"] = df["UF"].apply(lambda x: UF_SIGLA.get(x, x))

    db.close()

    xlsx_bytes = _build_excel(df, chart_title,
                              modalidade, fab, mod, uf, mun, ano_ini, ano_fim)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    return dcc.send_bytes(xlsx_bytes, filename=f"frota_motos_{slug}_{ts}.xlsx")


# ── Reset filters ──────────────────────────────────────────────────────────────
@app.callback(
    Output("f-modal",   "value"),
    Output("f-fab",     "value"),
    Output("f-uf",      "value"),
    Output("f-ano-ini", "value"),
    Output("f-ano-fim", "value"),
    Input("btn-reset", "n_clicks"),
    prevent_initial_call=True,
)
def reset_filters(_):
    return [], None, [], 1970, 2026


# ── Expand modal ───────────────────────────────────────────────────────────────
_MODAL_SHOW = {
    "display": "flex", "position": "fixed", "top": 0, "left": 0,
    "width": "100vw", "height": "100vh",
    "background": "rgba(27,73,101,0.6)",
    "zIndex": 999, "justifyContent": "center", "alignItems": "center",
    "backdropFilter": "blur(8px)",
}
_MODAL_HIDE = {"display": "none"}

_EXPAND_TITLES = {
    "expand-g-fab":     "Top Fabricantes",
    "expand-g-modelos": "Top 15 Modelos",
    "expand-g-anos":    "Evolução por Ano",
    "expand-g-munis":   "Top 20 Municípios",
}

@app.callback(
    Output("modal-overlay", "style"),
    Output("modal-graph",   "figure"),
    Output("modal-title",   "children"),
    Input("expand-g-fab",     "n_clicks"),
    Input("expand-g-modelos", "n_clicks"),
    Input("expand-g-anos",    "n_clicks"),
    Input("expand-g-munis",   "n_clicks"),
    Input("modal-close",      "n_clicks"),
    State("g-fab",     "figure"),
    State("g-modelos", "figure"),
    State("g-anos",    "figure"),
    State("g-munis",   "figure"),
    prevent_initial_call=True,
)
def handle_modal(n1, n2, n3, n4, n_close,
                 fig_fab, fig_mod, fig_ano, fig_mun):
    tid = ctx.triggered_id
    if tid == "modal-close":
        return _MODAL_HIDE, {}, ""

    figs = {
        "expand-g-fab":     fig_fab,
        "expand-g-modelos": fig_mod,
        "expand-g-anos":    fig_ano,
        "expand-g-munis":   fig_mun,
    }
    fig = copy.deepcopy(figs.get(tid) or {})
    if fig and "layout" in fig:
        fig["layout"]["height"]   = None
        fig["layout"]["autosize"] = True

    return _MODAL_SHOW, fig, _EXPAND_TITLES.get(tid, "")


# ── Run ────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=False, host="127.0.0.1", port=8052)
